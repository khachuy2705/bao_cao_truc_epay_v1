#!/usr/bin/python
# -*- coding: utf-8 -*- 
# import xlrd
# import xlwt
# baocao=xlrd.open_workbook("baocao.xlsx")
# baocao = xlwt.Workbook()
# # list_sheet=baocao.sheets()
# sheet = baocao.add_sheet('test')
# baocao.save('output.xls')
# for i in list_sheet:
# 	print i.name.encode("utf-8")

# # from openpyxl import Workbook
import openpyxl
import time
import cx_Oracle

ip = '172.16.100.61'
port = 1521
SID = 'XE'
useridOra ='REPORTER'
passwdOra = 'Abd123123'
# connstrOra = '127.0.0.1'
dsn_tns = cx_Oracle.makedsn(ip, port, SID)
dsnStr= "(DESCRIPTION=(ADDRESS_LIST=(FAILOVER=on)(LOAD_BALANCE=on)(SOURCE_ROUTE=off)(ADDRESS=(PROTOCOL=TCP)(HOST=172.16.9.61)(PORT=1521))(ADDRESS=(PROTOCOL=TCP)(HOST=172.16.9.62)(PORT=1521))(ADDRESS=(PROTOCOL=TCP)(HOST=172.16.9.63)(PORT=1521)))(CONNECT_DATA=(SERVICE_NAME=PDB4)))"
connect_stringDB = cx_Oracle.connect(useridOra, passwdOra, dsn=dsnStr)
curs = connect_stringDB.cursor()

print "Mở file báo cáo"
start_time = time.time()
wb = openpyxl.load_workbook('baocao.xlsx')
print"Đã mở file: %s giây" % (time.time() - start_time)
# wb.create_sheet('Du_lieu_1')
# start_time = time.time()
# print "Lấy danh sách sheet trong file"
# print wb.get_sheet_names()
# print "Đã lấy danh sách: %s giây" % (time.time() - start_time)

ws = wb.get_sheet_by_name('Du_lieu_1')
# for cot in ['A','B', 'C', 'D', 'E']:
# for hang in [1,200]:
# 	cell="D"+str(hang)
# 	ws[cell]=0
	
# ws['A1']=0
# # # Data can be assigned directly to cells
# ws['A1'] = 42

# # # Rows can also be appended
# # ws.append([1, 2, 3])

# # # Python types will automatically be converted
# # import datetime
# # ws['A2'] = datetime.datetime.now()

# # # Save the file
# # wb.save("sample.xlsx")
# a=wb.get_sheet_names()
# for a in wb.get_sheet_names():
# 	print unicode(str(a), 'unicode-escape')
print "Bat dau lay du lieu lan 1"
start_time = time.time()
sql_cmd = """SELECT '1._CDV_da_check', provider, COUNT (*) CDV_checked
    FROM CDV_CHARGING_TRANS_CENTER
   WHERE     request_time >= TRUNC (SYSDATE)
         AND request_time <= SYSDATE - 5 / 1440
         AND status <> 99
         AND old_status = 99
         AND CHANNEL = 'CDV'
GROUP BY provider
UNION
  SELECT '2.San_luong_account_dua_len',
         provider,
         COUNT (b.orders_id) SAN_ACC_DUA_LEN
    FROM CDV_charging_account a, CDV_CHARGING_ORDERS b
   WHERE     a.orders_id = b.orders_id
         AND a.TIME_Update >= TRUNC (SYSDATE)
         AND a.time_update < SYSDATE + 1
         AND a.account_type IN (1, 2)
GROUP BY provider
UNION
  SELECT '3.San_luong_thanh_cong', provider, COUNT (b.orders_id) SL_THANH_CONG
    FROM CDV_charging_account a, CDV_CHARGING_ORDERS b
   WHERE     a.orders_id = b.orders_id
         AND a.time_update >= TRUNC (SYSDATE)
         AND a.time_update < SYSDATE + 1
         AND a.amount_charging_success <> 0
         AND a.account_type IN (1, 2)
GROUP BY provider
UNION
  SELECT '4.TONG_DON_HANG_CHUA_CHAY',
         provider,
         COUNT (*) TONG_DON_HANG_CHUA_CHAY
    FROM CDV_CHARGING_ACCOUNT a
   WHERE     a.time_update >= TRUNC (SYSDATE)
         AND a.time_update < SYSDATE + 1
         AND amount_charging_success = 0
         AND account_type IN (1, 2)
         AND charging_status = 1
GROUP BY provider
UNION
  SELECT '5.Tu_dung_don_hang', provider, COUNT (*) TU_DUNG_DON_HANG
    FROM CDV_CHARGING_ACCOUNT
   WHERE     time_update >= TRUNC (SYSDATE)
         AND time_update < TRUNC (SYSDATE + 1)
         AND amount_charging_success = 0
         AND account_type IN (1, 2)
         AND charging_status = 0
GROUP BY provider
UNION
  SELECT '6.TOng_pending_CDV', provider, COUNT (*) TONG_PENDING
    FROM CDV_CHARGING_TRANS_CENTER
   WHERE     request_time >= TRUNC (SYSDATE)
         AND request_time <= SYSDATE - 5 / 1440
         AND (status = 99 OR old_status = 99)
         AND CHANNEL = 'CDV'
GROUP BY provider
UNION
  SELECT '7.Tong_GD_CDV', provider, COUNT (*) TONG_GIAO_DICH
    FROM CDV_CHARGING_TRANS_CENTER
   WHERE     Request_time >= TRUNC (SYSDATE)
         AND request_time <= SYSDATE - 10 / 1440
         AND CHANNEL = 'CDV'
GROUP BY provider
UNION
  SELECT '1.TOPUP_TONG_GIAO_DICH', service_provider_code, COUNT (*)
    FROM direct_transactions
   WHERE     trans_date >= TRUNC (SYSDATE)
         AND trans_date <= SYSDATE - 20 / 1440
         AND service_provider_code IN ('VMS',
                                       'VMS_EZ',
                                       'VNP',
                                       'VNP_ELOAD',
                                       'VTT',
                                       'VNM',
                                       'BEE')
GROUP BY service_provider_code
UNION
  SELECT '1.TOPUP_TONG_THAT_BAI', service_provider_code, COUNT (*)
    FROM direct_transactions
   WHERE     trans_date >= TRUNC (SYSDATE)
         AND trans_date <= SYSDATE - 20 / 1440
         AND service_provider_code IN ('VMS',
                                       'VMS_EZ',
                                       'VNP',
                                       'VNP_ELOAD',
                                       'VTT',
                                       'VNM',
                                       'BEE')
         AND trans_status IN ('13',
                              '17',
                              '12',
                              '18',
                              '11',
                              '6',
                              '16',
                              '98',
                              '8',
                              '-11',
                              '7',
                              '15')
GROUP BY service_provider_code
UNION
  SELECT '1.TOPUP_PENDING', service_provider_code, COUNT (*)
    FROM direct_transactions
   WHERE     trans_date >= TRUNC (SYSDATE)
         AND trans_date <= SYSDATE - 20 / 1440
         AND (trans_status = 99 OR trans_status_old = 99)
         AND service_provider_code IN ('VMS',
                                       'VMS_EZ',
                                       'VNP',
                                       'VNP_ELOAD',
                                       'VTT',
                                       'VNM',
                                       'BEE')
GROUP BY service_provider_code
UNION
  SELECT '1.TOPUP_PENDING_DA_CHECK', service_provider_code, COUNT (*)
    FROM direct_transactions
   WHERE     trans_date >= TRUNC (SYSDATE)
         AND trans_date <= SYSDATE - 20 / 1440
         AND service_provider_code IN ('VMS',
                                       'VMS_EZ',
                                       'VNP',
                                       'VNP_ELOAD',
                                       'VTT',
                                       'VNM',
                                       'BEE')
         AND (trans_status <> 99 AND trans_status_old = 99)
GROUP BY service_provider_code
UNION
  SELECT '2.CHARGING_TONG_GIAO_DICH', telco_code, COUNT (*)
    FROM tbl_transaction_logs
   WHERE     strdate = TO_NUMBER (TO_CHAR (SYSDATE, 'yyyymmdd'))
         AND receivetime >= TRUNC (SYSDATE)
         AND receivetime <= SYSDATE - 5 / 1440
GROUP BY telco_code
UNION ALL
  SELECT '2.CHARGING_TONG_PENDING', telco_code, COUNT (*)
    FROM tbl_transaction_logs
   WHERE     strdate = TO_NUMBER (TO_CHAR (SYSDATE, 'yyyymmdd'))
         AND receivetime >= TRUNC (SYSDATE)
         AND receivetime <= SYSDATE - 5 / 1440
         AND (status = 99 OR old_status = 99)
GROUP BY telco_code
UNION ALL
  SELECT '2.CHARGING_TONG_DA_CHECK', telco_code, COUNT (*)
    FROM tbl_transaction_logs
   WHERE     strdate = TO_NUMBER (TO_CHAR (SYSDATE, 'yyyymmdd'))
         AND receivetime >= TRUNC (SYSDATE)
         AND receivetime <= SYSDATE - 5 / 1440
         AND status <> 99
         AND old_status = 99
GROUP BY telco_code
UNION
SELECT '3.MGCP', 'Itopup_MGC_TONG GD', COUNT (*)
  FROM direct_transactions
 WHERE     partner_code = 'MGC'
       AND trans_date >= TRUNC (SYSDATE)
       AND trans_date < SYSDATE + 1
UNION
SELECT '3.MGCP', 'Itopup_MGC_pending', COUNT (*)
  FROM direct_transactions
 WHERE     partner_code = 'MGC'
       AND trans_date >= TRUNC (SYSDATE)
       AND trans_date < SYSDATE + 1
       AND (trans_status = 99 OR trans_status_old = 99)
UNION
  SELECT '4.Charging_tongthatbai', telco_code, COUNT (*)
    FROM tbl_transactions
   WHERE     strdate = TO_NUMBER (TO_CHAR (SYSDATE, 'yyyymmdd'))
         AND telco_code IN ('VMS',
                            'VNP',
                            'VTT',
                            'MGC')
         AND status NOT IN (1, 99)
GROUP BY telco_code
UNION ALL
  SELECT '4.Charging_thatbai_telco_epay', telco_code, COUNT (*)
    FROM tbl_transactions
   WHERE     strdate = TO_NUMBER (TO_CHAR (SYSDATE, 'yyyymmdd'))
         AND telco_code IN ('VMS',
                            'VNP',
                            'VTT',
                            'MGC')
         AND status IN (0,
                        13,
                        10,
                        11)
GROUP BY telco_code
UNION ALL
  SELECT '4.Charging_thatbai_nguoisudung', telco_code, COUNT (*)
    FROM tbl_transactions
   WHERE     strdate = TO_NUMBER (TO_CHAR (SYSDATE, 'yyyymmdd'))
         AND telco_code IN ('VMS',
                            'VNP',
                            'VTT',
                            'MGC')
         AND status NOT IN (0,
                            13,
                            10,
                            11,
                            1,
                            99)
GROUP BY telco_code
UNION ALL
SELECT '3.MGCP', 'CHARGING_MGC_TONG_GD', COUNT (*)
  FROM TBL_TRANSACTIONS
 WHERE     username = 'HM102'
       AND strdate = TO_NUMBER (TO_CHAR (SYSDATE, 'yyyymmdd'))
UNION
SELECT '3.MGCP', 'CHARGING_MGC_TONG_PENDING', COUNT (*)
  FROM TBL_TRANSACTIONS
 WHERE     username = 'HM102'
       AND strdate = TO_NUMBER (TO_CHAR (SYSDATE, 'yyyymmdd'))
       AND (STATUS = '99' OR OLD_STATUS = '99')
UNION
SELECT '3.MGCP', 'CHARGING_MGC_TONG_DA_CHECK', COUNT (*)
  FROM TBL_TRANSACTIONS
 WHERE     username = 'HM102'
       AND strdate = TO_NUMBER (TO_CHAR (SYSDATE, 'yyyymmdd'))
       AND OLD_STATUS = '99'"""
ket_qua = curs.execute(sql_cmd)
print"Da lay xong du lieu lan 1: %s giây" % (time.time() - start_time)
# connect_stringDB.commit()
# connect_stringDB.close()
ws.append(["THAM_SO", "NHA_MANG", "GIA_TRI"])
for row in ket_qua:
        row_new=[]
        row=list(row)
        a=str(row[0])+str(row[1])
        row_new.append(a)
        for i in row:
                row_new.append(i)
	ws.append(row_new)

sql_cmd2="""SELECT 'TONG_THAT_BAI',
         partner_name,
         PROVIDER,
         COUNT (*) AS Tong
    FROM CDV_ITU_TOPUP_TRANSACTION
   WHERE     partner_name IN ('IRISMEDIA',
                              'GMOBTOPUP_01',
                              'ononpay',
                              'PAYOO1',
                              'THEGIOIDIDONG_2',
                              'VIMO2',
                              'cty_htc',
                              'tikivn',
                              'vtaepay',
                              'BKCDNG',
                              'Mservice',
                              'VTCONLINE',
                              'garenaved',
                              'tappvn')
         AND REQUEST_TIME >= TRUNC (SYSDATE)
         AND STATUS = 110
         AND provider IN ('VTT', 'VNP', 'VMS')
GROUP BY partner_name, PROVIDER
UNION ALL
  SELECT 'TONG_GD',
         partner_name,
         PROVIDER,
         COUNT (*) AS TONG_GD
    FROM CDV_ITU_TOPUP_TRANSACTION
   WHERE     partner_name IN ('IRISMEDIA',
                              'GMOBTOPUP_01',
                              'ononpay',
                              'PAYOO1',
                              'THEGIOIDIDONG_2',
                              'VIMO2',
                              'cty_htc',
                              'tikivn',
                              'vtaepay',
                              'BKCDNG',
                              'Mservice',
                              'VTCONLINE',
                              'garenaved',
                              'tappvn')
         AND REQUEST_TIME >= TRUNC (SYSDATE)
         AND provider IN ('VTT', 'VNP', 'VMS')
GROUP BY partner_name, PROVIDER"""


# wb.create_sheet('Du_lieu_2')
ws = wb.get_sheet_by_name('Du_lieu_2')
# # for cot in ['A','B', 'C', 'D', 'E']:
for hang in [1,200,1]:
	cell='E'+str(hang)
	ws[cell]=0
ws['A1']=0
ket_qua2 = curs.execute(sql_cmd2)
ws.append(["THAM_SO", "NHA_MANG", "GIA_TRI"])
for row in ket_qua2:
        row_new=[]
        row=list(row)
        a=str(row[0])+str(row[1])
        row_new.append(a)
        for i in row:
                row_new.append(i)
 #       row_new.append(row)
 #       print row_new
#        row_new=list(row_new)
	ws.append(row_new)
wb.save('baocao_result.xlsx')
curs.close()
connect_stringDB.close()
